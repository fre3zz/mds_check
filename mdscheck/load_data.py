import os

from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")

import django
django.setup()

from mdscheck.models import MdsModel, Decision


data = load_workbook('data.xlsx')
sheet = data['Лист1']
data_dict = dict()
# Создаем словарь вида {x (номер кейса, не айди): ['y1' - экспертные данные,'y2','y3']
for row in range(2, 250):
    number = int(sheet.cell(row=row, column=1).value)

    if number:
        cd13_cd11b = ''
        cd16_cd13 = ''
        cd16_cd11b = ''
        if sheet.cell(row=row, column=2).value == 0:
            cd16_cd13 = 'neg'
        else:
            cd16_cd13 = 'pos'
        if sheet.cell(row=row, column=3).value == 0:
            cd16_cd11b = 'neg'
        else:
            cd16_cd11b = 'pos'
        if sheet.cell(row=row, column=4).value == 0:
            cd13_cd11b = 'neg'
        else:
            cd13_cd11b = 'pos'

        data_dict[number] = [cd13_cd11b, cd16_cd13, cd16_cd11b]

mds_cases = MdsModel.objects.all()

for mds_case in mds_cases:
    if mds_case.number in data_dict.keys():
        images = mds_case.images.all()
        for image in images:
            results = data_dict[mds_case.number]
            # выбор по индексу - image.number на один опережает индекс в словаре data
            decision = results[int(image.name)-1]
            new_decision, created = Decision.objects.get_or_create(
                image=image,
                is_expert=True
            )
            print(new_decision)
            new_decision.decision = decision
            new_decision.responder = 'exp'
            new_decision.responder_email = 'k0jl9ih@yandex.ru'
            new_decision.save()
